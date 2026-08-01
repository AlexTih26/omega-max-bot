"""Экспорт Таксимо в Excel."""

from __future__ import annotations

import io
from datetime import date
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

from taksimo_store import is_placed, list_registry_rows, list_sessions, yard_map

EXPORT_DIR = Path(__file__).resolve().parent.parent / "docs" / "registry" / "exports"


def _header_row(ws, titles: list[str]) -> None:
    ws.append(titles)
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)


def build_registry_workbook(
    *,
    date_from: str | None = None,
    date_to: str | None = None,
) -> bytes:
    rows = list_registry_rows(date_from=date_from, date_to=date_to)
    wb = Workbook()
    ws = wb.active
    ws.title = "Реестр"
    _header_row(
        ws,
        [
            "Дата",
            "Авто ГРЗ",
            "Водитель",
            "№ ТРН",
            "Буква",
            "№ блока",
            "№ ВАГОНА",
            "Место X/Y",
            "Пометка",
            "Кран с",
            "Кран до",
            "Кран мин",
            "Стропальщики ₽",
            "Такси ₽",
            "Примечания",
        ],
    )
    for r in rows:
        place = f"{r['pos_x']}/{r['pos_y']}"
        if r.get("suffix"):
            place += r["suffix"]
        ws.append(
            [
                r.get("unload_date"),
                r.get("vehicle_plate") or "",
                r.get("driver") or "",
                r.get("trn") or "",
                r.get("letter"),
                r.get("number"),
                r.get("wagon_number") or "",
                place,
                r.get("suffix") or "",
                r.get("crane_start") or "",
                r.get("crane_end") or "",
                r.get("crane_minutes") or "",
                r.get("riggers_pay") or "",
                r.get("taxi_pay") or "",
                r.get("session_notes") or "",
            ]
        )

    ws2 = wb.create_sheet("Журнал выгрузок")
    _header_row(
        ws2,
        [
            "Дата",
            "ТРН",
            "Авто",
            "Водитель",
            "Кран с",
            "Кран до",
            "Минут",
            "Стропальщики ₽",
            "Такси ₽",
            "Плит",
            "№ вагона",
            "Примечания",
        ],
    )
    for s in list_sessions(limit=500):
        if date_from and s["unload_date"] < date_from:
            continue
        if date_to and s["unload_date"] > date_to:
            continue
        ws2.append(
            [
                s["unload_date"],
                s.get("trn"),
                s.get("vehicle_plate") or "",
                s.get("driver") or "",
                s.get("crane_start"),
                s.get("crane_end"),
                s.get("crane_minutes"),
                s.get("riggers_pay"),
                s.get("taxi_pay"),
                s.get("slab_count", 0),
                s.get("wagon_numbers") or "",
                s.get("notes") or "",
            ]
        )

    ws3 = wb.create_sheet("Реестр бух")
    _header_row(
        ws3,
        [
            "№",
            "Дата и время выгрузки",
            "АВТО ГРЗ",
            "ВОДИТЕЛЬ ФИО",
            "№ ТРН ДАТА",
            "БУКВА БЛОКА",
            "№БЛОКА",
            "№ ВАГОНА",
            "ДАТА ПОГРУЗКИ ПВ",
            "ПЛОЩАДКА",
            "склад место гориз",
            "склад место глубина",
        ],
    )
    last_session: int | None = None
    reg_no = 0
    for r in rows:
        sid = r.get("session_id")
        if sid != last_session:
            reg_no += 1
            last_session = sid
            show_no = reg_no
        else:
            show_no = ""
        dt_show = r.get("unload_datetime") or r.get("unload_date") or ""
        block_label = f"{r.get('letter', '')} {r.get('number', '')}".strip()
        px = r["pos_x"] if is_placed(r["pos_x"], r["pos_y"]) else ""
        py = r["pos_y"] if is_placed(r["pos_x"], r["pos_y"]) else ""
        ws3.append(
            [
                show_no,
                dt_show,
                r.get("vehicle_plate") or "",
                r.get("driver") or "",
                r.get("trn") or "",
                r.get("letter"),
                block_label,
                r.get("wagon_number") or "",
                r.get("loading_date") or "",
                r.get("platform_zone") or "ХРАНЕНИЯ",
                px,
                py,
            ]
        )

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_yard_workbook() -> bytes:
    yard = yard_map()
    wb = Workbook()
    ws = wb.active
    ws.title = "Площадка"
    _header_row(
        ws,
        ["X", "Y", "Место", "Буква", "№ блока", "Пометка", "Дата выгрузки", "Авто", "Водитель"],
    )
    for _key, items in sorted(
        yard["cells"].items(),
        key=lambda kv: (int(kv[0].split("/")[1]), int(kv[0].split("/")[0])),
    ):
        for s in items:
            ws.append(
                [
                    s["pos_x"],
                    s["pos_y"],
                    s["place"] + (s.get("suffix") or ""),
                    s["letter"],
                    s["number"],
                    s.get("suffix") or "",
                    s.get("unload_date") or "",
                    s.get("vehicle_plate") or "",
                    s.get("driver") or "",
                ]
            )
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def save_registry_copy(
    data: bytes,
    *,
    date_from: str | None = None,
    date_to: str | None = None,
) -> Path:
    EXPORT_DIR.mkdir(parents=True, exist_ok=True)
    today = date.today().isoformat()
    if date_from and date_to and date_from != date_to:
        stem = f"reestr_{date_from}_{date_to}"
    elif date_from:
        stem = f"reestr_{date_from}"
    else:
        stem = f"reestr_{today}"
    path = EXPORT_DIR / f"{stem}.xlsx"
    path.write_bytes(data)
    return path
